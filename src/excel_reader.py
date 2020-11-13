import json
import os
import re
import sys
from openpyxl import load_workbook
from src.enums_and_constants import Colors, PARAMETER_PATTERN
from src.utility_functions import converter, calculate_the_value

# finding_parameters_in_each_section


def finding_parameters(sheet, data, section="BASIC_PARAMETERS",
                       start_index=0, end_index=100, formula_or_para="values",
                       color_param1=Colors.PARAMETER_COLOR.value,
                       color_param2=Colors.VALUE_COLOR.value):
    for row in sheet.iter_rows(min_row=start_index, max_row=end_index):
        for index, cell in enumerate(row):
            if cell.value is not None and \
                cell.font.color is not None and \
                    str(cell.font.color.index) == color_param1:
                param_value = cell.value
                for inner_cell in row[index:]:
                    if inner_cell.value != None and\
                        inner_cell.font.color is not None and \
                            inner_cell.font.color.index == color_param2:
                        value = converter(inner_cell.value)
                        if formula_or_para != "formula":
                            data[section][formula_or_para][inner_cell.coordinate] = {
                                'value': calculate_the_value(value, sheet) if re.match(PARAMETER_PATTERN, value) != None else value,
                                'name': param_value,
                            }
                        else:
                            data[section][formula_or_para][inner_cell.coordinate] = {
                                'value': value,
                                'name': param_value,
                                'calculated_value': calculate_the_value(value, sheet)
                            }
                        break
                del param_value
                break

# finding sections


def finding_sections(sheet, data):
    for index, row in enumerate(sheet.rows):
        if re.search(r"\d+\w*\)", str(row[0].value)):
            data[str(row[1].value)] = {
                'position': row[1].coordinate,
                'start_index': index,
                "values": {},
                "formula": {}
            }
            continue

# find each section


def parse_sections(sheet, data, formula_or_param='values',
                   color_param1=Colors.PARAMETER_COLOR.value,
                   color_param2=Colors.VALUE_COLOR.value):
    keys = list(data.keys())
    for index, key in enumerate(keys[:len(keys) - 1]):
        finding_parameters(sheet=sheet,
                           data=data,
                           section=key,
                           start_index=int(data[key]['start_index']),
                           end_index=int(data[keys[index + 1]]['start_index']),
                           formula_or_para=formula_or_param,
                           color_param1=color_param1,
                           color_param2=color_param2)

    finding_parameters(sheet=sheet,
                       data=data,
                       section=keys[len(keys) - 1],
                       start_index=int(
                           data[keys[len(keys) - 1]]['start_index']),
                       end_index=int(len(list(sheet.rows))),
                       formula_or_para=formula_or_param,
                       color_param1=color_param1,
                       color_param2=color_param2)


def initialing_parse_excel(file_path=""):
    if len(file_path) == 0 or file_path == "":
        raise Exception("File Not Specified")

    if not os.path.exists(file_path):
        raise Exception("File Path Does Not Exists.")

    workbook, sheet, data = None, None, None
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        data = dict(BASIC_PARAMETERS=dict(
            {'position': 'A1', 'start_index': 0, 'values': {}, 'formula': {}}))

        finding_sections(sheet=sheet, data=data)
        # it will find parameters
        parse_sections(sheet, data)
        # this will find formulas
        parse_sections(sheet, data,
                       formula_or_param='formula',
                       color_param1=Colors.DRIVED_COLOR.value,
                       color_param2=Colors.FORMULA_COLOR.value)
    except Exception as ex:
        print(ex)
        raise Exception(ex.args)
    else:
        return (sheet, data, workbook)

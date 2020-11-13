# class Dialog(QMainWindow):
#     NumGridRows = 3
#     NumButtons = 4

#     def __init__(self):
#         super(Dialog, self).__init__()
#         self.setMinimumWidth(1000)

#         self.createMenu()
#         self.createHorizontalGroupBox()
#         self.createGridGroupBox()
#         self.createFormGroupBox()

#         bigEditor = QTextEdit()
#         bigEditor.setPlainText("This widget takes up all the remaining space "
#                 "in the top-level layout.")

#         # buttonBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)

#         # buttonBox.accepted.connect(self.accept)
#         # buttonBox.rejected.connect(self.reject)

#         mainLayout = QVBoxLayout()
#         # mainLayout.setMenuBar(self.menuBar)
#         # mainLayout.addWidget(self.horizontalGroupBox)
#         # mainLayout.addWidget(self.gridGroupBox)
#         # mainLayout.addWidget(self.formGroupBox)
#         # mainLayout.addWidget(bigEditor)
#         # mainLayout.addWidget(buttonBox)
#         self.setLayout(mainLayout)

#         toolbar = QToolBar("Main Toolbar")
#         # self.setToolTip()

#         self.setWindowTitle("Basic Layouts")

#     def createMenu(self):
#         self.menuBar = QMenuBar()

#         self.fileMenu = QMenu("&File", self)
#         self.exitAction = self.fileMenu.addAction("E&xit")
#         self.menuBar.addMenu(self.fileMenu)

#         # self.exitAction.triggered.connect(self.accept)

#     def createHorizontalGroupBox(self):
#         self.horizontalGroupBox = QGroupBox("Horizontal layout")
#         layout = QHBoxLayout()

#         for i in range(Dialog.NumButtons):
#             button = QPushButton("Button %d" % (i + 1))
#             layout.addWidget(button)

#         self.horizontalGroupBox.setLayout(layout)

#     def createGridGroupBox(self):
#         self.gridGroupBox = QGroupBox("Grid layout")
#         layout = QGridLayout()

#         for i in range(Dialog.NumGridRows):
#             label = QLabel("Line %d:" % (i + 1))
#             lineEdit = QLineEdit()
#             layout.addWidget(label, i + 1, 0)
#             layout.addWidget(lineEdit, i + 1, 1)

#         self.smallEditor = QTextEdit()
#         self.smallEditor.setPlainText("This widget takes up about two thirds "
#                 "of the grid layout.")

#         layout.addWidget(self.smallEditor, 0, 2, 4, 1)

#         layout.setColumnStretch(1, 10)
#         layout.setColumnStretch(2, 20)
#         self.gridGroupBox.setLayout(layout)

import re
from openpyxl.reader.excel import load_workbook
import json
import sys

FORMULA_COLOR = "FFFF0000"
DRIVED_COLOR = "5"

VAR_PATTERN = r'[A-Z]+\d{1,5}'

CONVENTIONS = {
    'MAX': 'max',
    'MIN': 'min',
    '=': '',
    'SQRT': 'sqrt',
    'SUM': 'sum',
    '^': '**'
}


def converter(val):
    pattern = '|'.join(map(re.escape, CONVENTIONS.keys()))
    return re.sub(pattern, lambda x: CONVENTIONS[x.group()], str(val))


def detecting_formula(sheet, data, starting_index, end_index, section):
    data[section]['formula'] = {}
    
    for row in sheet.iter_rows(min_row=starting_index, max_row=end_index):
        for index, cell in enumerate(row):
            if cell.value is not None and \
                cell.font.color is not None and \
                    str(cell.font.color.index) == FORMULA_COLOR:
                
                flag = False
                formula_value = cell.value
                for inner_cell in row[:index]:
                    if inner_cell.value != None and\
                        inner_cell.font.color is not None and \
                            str(inner_cell.font.color.index) == DRIVED_COLOR:
                        # print(formula_value)
                        data[section]['formula'][inner_cell.value] = {
                            'value': converter(val=formula_value),
                            'position': cell.coordinate
                        }
                        flag = True
                        break
                if not flag:
                    data[section]['formula'][cell.coordinate] = {
                        'value': converter(val=formula_value),
                        'position': cell.coordinate
                    }
                del formula_value
                break


def searching_in_variables_or_formula(value, data: dict, val_or_formula):
    for section in data.keys():
        if isinstance(data[section][val_or_formula], dict):
            for key in data[section][val_or_formula].keys():
                if data[section][val_or_formula][key]['position'] == value:
                    return True, section, key
    return False, None, None


def replacements(value, data: dict):
    # print(re.sub(VAR_PATTERN, lambda x: print(x.group()), str(value)))
    variables = re.findall(VAR_PATTERN, value)
    main_variables = {}
    print('variables is : ', variables)
    
    for var in variables:
        Found, sec, k = searching_in_variables_or_formula(var, data, 'values')
        if Found:
            main_variables[var] = {
                'key': k,
                'section': sec
            }
        else:
            formula_found, section, key = searching_in_variables_or_formula(var, data, 'formula')
            if formula_found:
                ret = replacements(data[section]['formula'][key]['value'], data)
                if ret != None:
                    main_variables = {
                        **main_variables,
                        var: {
                            **ret
                        }
                    }
    
    if len(main_variables) == len(variables):
        return main_variables
            


def parsing_formula_into_variables(data: dict):
    for section in data.keys():
        for cell in data[section]['formula'].values():
            val = cell['value']
            if cell.get('drived_formula', None) == None:
                cell['drived_formula'] = replacements(val, data)
            # print(cell['drived_formula'])
    
    f = open('formula.json', "w")
    f.write(json.dumps(data))
    f.close()
    

def main():
    workbook = load_workbook("test.xlsx")
    sheet = workbook.active

    f = open("parameters.json") 

    sections = json.loads(f.read())
    f.close()
    # parsing_formula_into_variables(sections)
    # sheet["A1"] = 'alli'
    # sys.exit(0)
    keys = list(sections.keys())

    for index, key in enumerate(keys):
        if index == len(keys) - 1:
            break
        detecting_formula(sheet,
                          sections,
                          int(sections[key]["start_index"]),
                          int(sections[keys[index + 1]]["start_index"]),
                          key)

    detecting_formula(sheet,
                      sections,
                      sections[keys[len(keys) - 1]]["start_index"],
                      len(list(sheet.rows)),
                      keys[-1])

    f = open('formula.json', "w")
    f.write(json.dumps(sections))
    f.close()


if __name__ == '__main__':
    main()
